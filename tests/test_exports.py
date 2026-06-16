from __future__ import annotations

from openpyxl import load_workbook

from construction_maintenance import repositories as repo
from construction_maintenance.services.exports import build_people_workbook


def test_build_people_workbook(app, tmp_path):
    with app.app_context():
        repo.create_person(
            {
                "name": "王小明",
                "id_number": "410000199001011234",
                "gender": "男",
                "birth_date": "1990-01-01",
                "age": 36,
                "phone": "13800000000",
                "address": "河南省郑州市",
                "job_type": "普工",
                "bank_card": "6222000000000000",
                "bank_name": "建设银行",
                "entry_date": "2026-05-29",
                "notes": "",
            }
        )
        output = tmp_path / "people.xlsx"
        build_people_workbook(output)

    workbook = load_workbook(output)
    sheet = workbook.active
    assert sheet["A1"].value == "姓名"
    assert sheet["A2"].value == "王小明"
    assert sheet["B2"].value == "410000199001011234"


def test_build_qualification_workbook(app, tmp_path):
    with app.app_context():
        company = repo.get_main_company()
        repo.create_qualification(
            {
                "company_id": company["id"],
                "name": "建筑业企业资质",
                "certificate_no": "D300000",
                "issue_date": "2026-01-01",
                "expiry_date": "2029-01-01",
                "is_long_term": 0,
                "attachment_path": "",
                "notes": "",
            }
        )
        output = tmp_path / "qualifications.xlsx"
        from construction_maintenance.services.exports import build_qualification_workbook
        build_qualification_workbook(output)

    workbook = load_workbook(output)
    sheet = workbook.active
    assert sheet["A1"].value == "公司"
    assert sheet["B2"].value == "建筑业企业资质"


def test_build_attendance_workbook(app, tmp_path):
    with app.app_context():
        # 1. 创建测试人员
        person_id = repo.create_person(
            {
                "name": "李小华",
                "id_number": "410000199001019999",
                "is_attendance": 1,
            }
        )
        # 2. 保存一笔考勤
        repo.save_attendance(person_id, "2026-06-01", "白班")
        repo.save_attendance(person_id, "2026-06-02", "请假")
        
        # 3. 导出考勤表到临时文件
        from construction_maintenance.services.exports import build_attendance_workbook
        output = tmp_path / "attendance.xlsx"
        workbook = build_attendance_workbook("2026-06")
        workbook.save(output)
        
    # 4. 验证导出的数据
    workbook = load_workbook(output)
    sheet = workbook.active
    
    # 检查表头
    assert sheet["A1"].value == "姓名"
    assert sheet["B1"].value == "身份证号"
    assert sheet["C1"].value == "1日"
    
    # 检查数据
    assert sheet["A2"].value == "李小华"
    assert sheet["B2"].value == "410000199001019999"
    # C2 (1日) 应该是 "上"
    assert sheet["C2"].value == "上"
    # D2 (2日) 应该是 "假"
    assert sheet["D2"].value == "假"


def test_build_contract_workbook(app, tmp_path):
    with app.app_context():
        # 先创建一个测试项目
        main_company = repo.get_main_company()
        proj_id = repo.create_project({
            "company_id": main_company["id"],
            "name": "导出测试项目",
            "status": "进行中"
        })
        
        # 创建一个测试合同
        repo.create_contract({
            "project_id": proj_id,
            "name": "导出测试合同",
            "contract_type": "材料商合同",
            "notes": "导出备注"
        })
        
        from construction_maintenance.services.exports import build_contract_workbook
        output = tmp_path / "contracts.xlsx"
        build_contract_workbook(output)
        
    workbook = load_workbook(output)
    sheet = workbook.active
    assert sheet.cell(1, 1).value == "合同ID"
    assert sheet.cell(1, 2).value == "归属项目"
    assert sheet.cell(1, 3).value == "合同名称"
    assert sheet.cell(1, 4).value == "合同分类"
    assert sheet.cell(1, 5).value == "备注"
    assert sheet.cell(1, 6).value == "创建时间"
    
    # 检查写入的数据
    assert sheet.cell(2, 2).value == "导出测试项目"
    assert sheet.cell(2, 3).value == "导出测试合同"
    assert sheet.cell(2, 4).value == "材料商合同"
    assert sheet.cell(2, 5).value == "导出备注"


