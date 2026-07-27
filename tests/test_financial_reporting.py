from io import BytesIO
from concurrent.futures import ThreadPoolExecutor
from threading import Event

from openpyxl import load_workbook

from construction_maintenance import repositories as repo
from construction_maintenance.db import get_db
from construction_maintenance.services.dashboard import build_dashboard
from construction_maintenance.services.exports import build_project_ledger_workbook
from construction_maintenance.services import exports as export_service


EXPECTED_HEADERS = [
    "记录编号", "日期", "项目", "收支类型", "一级分类", "二级分类",
    "事项摘要", "金额", "经办/垫付人", "付款状态", "付款日期",
    "支付/报销说明", "复核状态", "分类置信度", "来源文件",
    "来源工作表", "原始行号", "作废状态",
]


def _leaf_id(name: str) -> int:
    return int(get_db().execute(
        "select id from expense_categories where name = ? and parent_id is not null",
        (name,),
    ).fetchone()["id"])


def seed_financial_entries():
    company = repo.get_main_company()
    project_id = repo.create_project({
        "company_id": company["id"],
        "name": "财务口径测试项目",
    })
    rows = [
        ("支出", "材料运输费", 1000),
        ("冲减支出", "材料运输费", 100),
        ("收入", "废料处置收入", 50),
        ("资金往来", "备用金", 200),
    ]
    for index, (transaction_type, category, amount) in enumerate(rows, start=1):
        repo.create_voucher({
            "project_id": project_id,
            "voucher_date": "2026-07-01",
            "transaction_type": transaction_type,
            "category_id": _leaf_id(category),
            "amount": amount,
            "notes": "测试事项",
            "source_record_id": f"REPORT-{index}",
            "source_filename": "=HYPERLINK(\"https://invalid\")" if index == 1 else "source.xls",
            "source_sheet": "汇总",
            "source_row": index,
            "payment_status": "未支付",
        })
    return project_id


def test_dashboard_uses_transaction_aware_financial_totals(app):
    with app.app_context():
        seed_financial_entries()
        dashboard = build_dashboard()

    assert dashboard["expense"] == 1000
    assert dashboard["expense_reduction"] == 100
    assert dashboard["net_expense"] == 900
    assert dashboard["income"] == 50
    assert dashboard["fund_transfer"] == 200


def test_project_ledger_export_contains_structured_fields_and_escapes_formulas(
    app, tmp_path
):
    with app.app_context():
        seed_financial_entries()
        output = build_project_ledger_workbook(tmp_path / "ledger.xlsx")

    sheet = load_workbook(output).active
    assert [cell.value for cell in sheet[1]] == EXPECTED_HEADERS
    assert "'=HYPERLINK(\"https://invalid\")" in [
        sheet.cell(row=row, column=15).value for row in range(2, sheet.max_row + 1)
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == f"A1:R{sheet.max_row}"


def test_filtered_ledger_export_includes_all_matches_across_pages(client, app):
    with app.app_context():
        company = repo.get_main_company()
        project_id = repo.create_project({
            "company_id": company["id"],
            "name": "筛选导出测试",
        })
        category_id = _leaf_id("五金辅材及工具")
        income_category_id = _leaf_id("废料处置收入")
        for index in range(30):
            repo.create_voucher({
                "project_id": project_id,
                "voucher_date": "2026-07-01",
                "transaction_type": "支出",
                "category_id": category_id,
                "amount": index + 1,
                "notes": f"EXPORT_MATCH_{index:02d}",
                "payment_status": "未支付",
            })
        repo.create_voucher({
            "project_id": project_id,
            "voucher_date": "2026-07-01",
            "transaction_type": "收入",
            "category_id": income_category_id,
            "amount": 500,
            "notes": "EXPORT_NONMATCH",
            "payment_status": "已支付/已报销",
        })

    response = client.get(
        f"/exports/project-ledger?project_id={project_id}&transaction_type=支出"
    )

    assert response.status_code == 200
    sheet = load_workbook(BytesIO(response.data)).active
    notes = [sheet.cell(row=row, column=7).value for row in range(2, sheet.max_row + 1)]
    assert len(notes) == 30
    assert "EXPORT_MATCH_00" in notes
    assert "EXPORT_MATCH_29" in notes
    assert "EXPORT_NONMATCH" not in notes


def test_concurrent_filtered_ledger_exports_are_request_isolated(
    app, monkeypatch
):
    with app.app_context():
        company = repo.get_main_company()
        category_id = _leaf_id("五金辅材及工具")
        project_ids = {}
        for marker in ("ONLY_A", "ONLY_B"):
            project_id = repo.create_project({
                "company_id": company["id"],
                "name": f"并发导出-{marker}",
            })
            repo.create_voucher({
                "project_id": project_id,
                "voucher_date": "2026-07-01",
                "transaction_type": "支出",
                "category_id": category_id,
                "amount": 100,
                "notes": marker,
                "payment_status": "未支付",
            })
            project_ids[marker] = project_id

    first_written = Event()
    second_written = Event()
    original_builder = export_service.build_project_ledger_workbook

    def coordinated_builder(target, project_id=None, **filters):
        output = original_builder(target, project_id=project_id, **filters)
        if project_id == project_ids["ONLY_A"]:
            first_written.set()
            assert second_written.wait(timeout=5)
        else:
            assert first_written.wait(timeout=5)
            second_written.set()
        return output

    monkeypatch.setattr(
        export_service,
        "build_project_ledger_workbook",
        coordinated_builder,
    )

    def download(project_id):
        with app.test_client() as thread_client:
            response = thread_client.get(
                f"/exports/project-ledger?project_id={project_id}"
            )
            assert response.status_code == 200
            return response.data

    with ThreadPoolExecutor(max_workers=2) as executor:
        first = executor.submit(download, project_ids["ONLY_A"])
        assert first_written.wait(timeout=5)
        second = executor.submit(download, project_ids["ONLY_B"])
        payloads = {
            "ONLY_A": first.result(timeout=5),
            "ONLY_B": second.result(timeout=5),
        }

    for marker, payload in payloads.items():
        sheet = load_workbook(BytesIO(payload)).active
        notes = {
            sheet.cell(row=row, column=7).value
            for row in range(2, sheet.max_row + 1)
        }
        assert notes == {marker}
