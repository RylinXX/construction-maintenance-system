from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from construction_maintenance.services.ledger_import import parse_ledger_source


DETAIL_HEADERS = [
    "记录编号", "项目名称", "发生日期", "年月", "收支类型", "一级分类",
    "二级分类", "事项摘要", "金额（元）", "净支出（元）", "收入金额（元）",
    "经办/垫付人", "支付状态", "支付日期", "支付/报销说明",
    "备注（原始用途/购买原因）", "审核标记", "分类置信度", "来源文件",
    "来源工作表", "原始行号", "原始金额（元）",
]
PENDING_HEADERS = [
    "记录类型", "项目名称", "发生日期", "事项/原始用途", "金额（元）",
    "当前一级分类", "当前二级分类", "待确认问题", "经办/垫付人",
    "支付说明", "来源文件", "来源工作表", "原始行号",
]


def make_ledger_zip(tmp_path: Path) -> Path:
    workbook = Workbook()
    detail = workbook.active
    detail.title = "费用明细"
    detail.append(["测试项目｜费用明细"])
    detail.append([])
    detail.append(DETAIL_HEADERS)
    detail.append([
        "测试-20260701-0001", "测试项目", "2026-07-01", "2026-07", "支出",
        "材料费", "五金辅材及工具", "工具采购", 100, 100, 0, "张三",
        "已支付/已报销", "2026-07-02", "现金支付", "原始备注", "", "高",
        "测试源.xls", "汇总", 8, 100,
    ])
    pending = workbook.create_sheet("待确认清单")
    pending.append(["测试项目｜待确认清单"])
    pending.append([])
    pending.append(PENDING_HEADERS)
    pending.append([
        "有金额待复核", "测试项目", "2026-07-01", "工具采购", 100,
        "材料费", "五金辅材及工具", "分类待确认", "张三", "现金支付",
        "测试源.xls", "汇总", 8,
    ])
    pending.append([
        "缺少金额", "测试项目", "2026-07-03", "吊车一个班", None,
        "机械设备费", "起重及装卸设备", "缺少金额/仅有业务数量记录",
        "李四", "", "测试源.xls", "汇总", 9,
    ])
    xlsx_path = tmp_path / "测试项目_标准费用账套.xlsx"
    workbook.save(xlsx_path)
    zip_path = tmp_path / "账套.zip"
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as archive:
        archive.write(xlsx_path, xlsx_path.name)
    return zip_path


def test_parser_marks_review_and_separates_missing_amount(tmp_path):
    preview = parse_ledger_source(make_ledger_zip(tmp_path))

    assert preview.project_count == 1
    assert len(preview.entries) == 1
    assert preview.entries[0].review_status == "待复核"
    assert len(preview.pending_items) == 1
    assert preview.pending_items[0].summary == "吊车一个班"
    assert preview.totals["expense"] == 100


def mutate_workbook_in_zip(zip_path: Path, tmp_path: Path, mutation) -> Path:
    with ZipFile(zip_path) as archive:
        member_name = archive.namelist()[0]
        workbook_path = tmp_path / "mutated.xlsx"
        workbook_path.write_bytes(archive.read(member_name))
    workbook = load_workbook(workbook_path)
    mutation(workbook)
    workbook.save(workbook_path)
    mutated_zip = tmp_path / "mutated.zip"
    with ZipFile(mutated_zip, "w", ZIP_DEFLATED) as archive:
        archive.write(workbook_path, member_name)
    return mutated_zip


def test_parser_rejects_unknown_category(tmp_path):
    source = make_ledger_zip(tmp_path)
    invalid = mutate_workbook_in_zip(
        source, tmp_path, lambda workbook: setattr(
            workbook["费用明细"]["G4"], "value", "不存在分类"
        )
    )
    with pytest.raises(ValueError, match="分类与收支类型不匹配"):
        parse_ledger_source(invalid)


def test_parser_rejects_missing_sheet(tmp_path):
    source = make_ledger_zip(tmp_path)
    invalid = mutate_workbook_in_zip(
        source, tmp_path, lambda workbook: workbook.remove(workbook["费用明细"])
    )
    with pytest.raises(ValueError, match="必须包含费用明细和待确认清单"):
        parse_ledger_source(invalid)


def test_parser_rejects_zip_path_traversal(tmp_path):
    source = tmp_path / "unsafe.zip"
    with ZipFile(source, "w", ZIP_DEFLATED) as archive:
        archive.writestr("../账套.xlsx", b"invalid")
    with pytest.raises(ValueError, match="不安全的压缩包路径"):
        parse_ledger_source(source)


def append_detail_copy(workbook, *, project_name=None):
    sheet = workbook["费用明细"]
    values = [cell.value for cell in sheet[4]]
    if project_name is not None:
        values[1] = project_name
        values[0] = "OTHER-20260701-0002"
    sheet.append(values)


def test_parser_rejects_mixed_projects(tmp_path):
    invalid = mutate_workbook_in_zip(
        make_ledger_zip(tmp_path), tmp_path,
        lambda workbook: append_detail_copy(workbook, project_name="另一个项目"),
    )
    with pytest.raises(ValueError, match="一个账套只能包含一个项目"):
        parse_ledger_source(invalid)


def test_parser_rejects_non_positive_amount(tmp_path):
    invalid = mutate_workbook_in_zip(
        make_ledger_zip(tmp_path), tmp_path,
        lambda workbook: setattr(workbook["费用明细"]["I4"], "value", 0),
    )
    with pytest.raises(ValueError, match="金额无效"):
        parse_ledger_source(invalid)


def test_parser_rejects_duplicate_record_id(tmp_path):
    invalid = mutate_workbook_in_zip(
        make_ledger_zip(tmp_path), tmp_path,
        lambda workbook: append_detail_copy(workbook),
    )
    with pytest.raises(ValueError, match="重复记录编号"):
        parse_ledger_source(invalid)
