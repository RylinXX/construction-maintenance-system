from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


from construction_maintenance import repositories as repo

PEOPLE_HEADERS = [
    "姓名",
    "身份证号",
    "性别",
    "出生日期",
    "年龄",
    "电话",
    "住址",
    "岗位/工种",
    "银行卡号",
    "开户行",
    "入职/进场日期",
    "备注",
]


def build_people_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "基础人员信息"
    sheet.append(PEOPLE_HEADERS)
    for person in repo.list_people():
        sheet.append(
            [
                person["name"],
                person["id_number"],
                person["gender"],
                person["birth_date"],
                person["age"],
                person["phone"],
                person["address"],
                person["job_type"],
                person["bank_card"],
                person["bank_name"],
                person["entry_date"],
                person["notes"],
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


QUALIFICATION_HEADERS = ["公司", "资质名称", "证书编号", "发证日期", "到期日期", "长期有效", "附件", "备注"]


def build_qualification_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "企业资质清单"
    sheet.append(QUALIFICATION_HEADERS)
    for item in repo.list_qualifications():
        sheet.append(
            [
                item["company_name"],
                item["name"],
                item["certificate_no"],
                item["issue_date"],
                item["expiry_date"],
                "是" if item["is_long_term"] else "否",
                item["attachment_path"],
                item["notes"],
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


PROJECT_LEDGER_HEADERS = ["日期", "项目", "类型", "金额", "备注", "附件", "录入人"]


def build_project_ledger_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "项目台账"
    sheet.append(PROJECT_LEDGER_HEADERS)
    for item in repo.list_vouchers():
        sheet.append(
            [
                item["voucher_date"],
                item["project_name"],
                item["voucher_type"],
                item["amount"],
                item["notes"],
                item["attachment_path"],
                item["entry_user"],
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def build_attendance_workbook(month: str, is_template: bool = False) -> Workbook:
    import calendar
    import datetime

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{month} 考勤表"
    
    # 启用网格线显示
    sheet.views.sheetView[0].showGridLines = True

    try:
        year, month_num = map(int, month.split("-"))
    except (ValueError, TypeError):
        now = datetime.datetime.now()
        year, month_num = now.year, now.month
        month = f"{year:04d}-{month_num:02d}"

    _, total_days = calendar.monthrange(year, month_num)

    # 1. 构造表头列
    headers = ["姓名", "身份证号"]
    for day in range(1, total_days + 1):
        headers.append(f"{day}日")
    headers.extend(["出勤天数", "请假天数"])

    sheet.append(headers)

    # 2. 设置表头样式（精致的淡蓝色背景、加粗、居中对齐、细灰框）
    header_fill = PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="0f172a")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="cbd5e1"),
        right=Side(style="thin", color="cbd5e1"),
        top=Side(style="thin", color="cbd5e1"),
        bottom=Side(style="thin", color="cbd5e1"),
    )

    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 3. 加载当前考勤名单
    people = repo.list_attendance_people()
    
    # 4. 加载当前月份的已有考勤数据
    attendance_dict = {}
    if not is_template:
        raw_attendance = repo.list_attendance_by_month(month)
        for record in raw_attendance:
            p_id = record["person_id"]
            date_str = record["work_date"]
            shift = record["shift_type"]
            if p_id not in attendance_dict:
                attendance_dict[p_id] = {}
            attendance_dict[p_id][date_str] = shift

    # 5. 循环写入数据行
    for r_idx, person in enumerate(people, start=2):
        row_data = [person["name"], person["id_number"]]
        day_count = 0
        leave_count = 0

        for d in range(1, total_days + 1):
            date_str = f"{month}-{d:02d}"
            shift = attendance_dict.get(person["id"], {}).get(date_str, None)
            if shift in ("白班", "夜班", "上班"):
                row_data.append("上")
                day_count += 1
            elif shift == "请假":
                row_data.append("假")
                leave_count += 1
            else:
                row_data.append("")

        row_data.extend([day_count, leave_count])
        sheet.append(row_data)

        # 为数据单元格应用精致的居中、边框及部分特殊标识字体样式
        for c_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = Font(name="微软雅黑", size=10)
            
            # 高亮标识白、夜、假
            val = cell.value
            if val == "白":
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="0284c7")  # 科技天蓝
            elif val == "夜":
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="1e40af")  # 深海蓝
            elif val == "假":
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="d97706")  # 橙黄色

    # 6. 自适应设置列宽
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or "")
            # 处理中文长度
            byte_len = len(val_str.encode("utf-8"))
            char_len = len(val_str)
            approx_len = char_len + (byte_len - char_len) // 2
            if approx_len > max_len:
                max_len = approx_len
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # 确保身份证号那列列宽较宽
    sheet.column_dimensions["B"].width = 22

    if not is_template:
        # 新增第二个 Sheet
        sal_sheet = workbook.create_sheet(title="工资结算与流水")
        sal_sheet.views.sheetView[0].showGridLines = True
        
        # 获取结算数据并分类
        summary_data = repo.get_salary_summary_by_month(month)
        long_term_data = [item for item in summary_data if item["salary_type"] in ["月薪", "年薪"]]
        daily_data = [item for item in summary_data if item["salary_type"] == "日薪"]
        
        # 样式定义
        title_font = Font(name="微软雅黑", size=12, bold=True, color="0f766e")
        header_fill = PatternFill(start_color="ecfdf5", end_color="ecfdf5", fill_type="solid")
        header_font = Font(name="微软雅黑", size=10, bold=True, color="0f172a")
        
        # 1. 写入长期员工工资结算表
        sal_sheet.cell(row=1, column=1, value="一、长期员工工资结算汇总表").font = title_font
        long_headers = ["姓名", "工种/岗位", "计薪方式", "薪资标准", "请假天数", "当月应发工资", "已预支金额", "已发放金额", "本月应补尾款"]
        
        # 格式化长期表头
        for col_idx, h in enumerate(long_headers, start=1):
            cell = sal_sheet.cell(row=3, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        current_row = 4
        for item in long_term_data:
            rate_str = f"¥{item['salary_rate']:.2f}/" + ("月" if item["salary_type"] == "月薪" else "年")
            att_str = f"{item['leave']}天"
            row_data = [
                item["name"],
                item["job_type"] or "技术/管理",
                item["salary_type"],
                rate_str,
                att_str,
                item["earnings"],
                item["advance"],
                item["payout"],
                item["balance"]
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = sal_sheet.cell(row=current_row, column=col_idx, value=val)
                cell.font = Font(name="微软雅黑", size=10)
                cell.border = thin_border
                if col_idx in [6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '"¥"#,##0.00'
                    if col_idx == 9:
                        if cell.value < 0:
                            cell.font = Font(name="微软雅黑", size=10, bold=True, color="ef4444")
                        elif cell.value > 0:
                            cell.font = Font(name="微软雅黑", size=10, bold=True, color="0f766e")
                else:
                    cell.alignment = center_align
            current_row += 1
            
        if not long_term_data:
            sal_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            empty_cell = sal_sheet.cell(row=current_row, column=1, value="当月无长期人员（月薪/年薪）结算数据")
            empty_cell.alignment = center_align
            empty_cell.font = Font(name="微软雅黑", size=10, italic=True, color="94a3b8")
            for col_idx in range(1, 10):
                sal_sheet.cell(row=current_row, column=col_idx).border = thin_border
            current_row += 1
            
        # 2. 空白隔断，写入日结员工工资结算表
        current_row += 2
        sal_sheet.cell(row=current_row, column=1, value="二、日结工人工资结算汇总表").font = title_font
        current_row += 1
        
        daily_headers = ["姓名", "工种/岗位", "计薪方式", "薪资标准", "实际出勤", "当月应发工资", "已预支金额", "已发放金额", "本月应补尾款"]
        for col_idx, h in enumerate(daily_headers, start=1):
            cell = sal_sheet.cell(row=current_row, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        current_row += 1
        for item in daily_data:
            rate_str = f"¥{item['salary_rate']:.2f}/天"
            att_str = f"{item['work_days']}天"
            row_data = [
                item["name"],
                item["job_type"] or "普工",
                item["salary_type"],
                rate_str,
                att_str,
                item["earnings"],
                item["advance"],
                item["payout"],
                item["balance"]
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = sal_sheet.cell(row=current_row, column=col_idx, value=val)
                cell.font = Font(name="微软雅黑", size=10)
                cell.border = thin_border
                if col_idx in [6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '"¥"#,##0.00'
                    if col_idx == 9:
                        if cell.value < 0:
                            cell.font = Font(name="微软雅黑", size=10, bold=True, color="ef4444")
                        elif cell.value > 0:
                            cell.font = Font(name="微软雅黑", size=10, bold=True, color="0f766e")
                else:
                    cell.alignment = center_align
            current_row += 1
            
        if not daily_data:
            sal_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            empty_cell = sal_sheet.cell(row=current_row, column=1, value="当月无日结工（日薪）结算数据")
            empty_cell.alignment = center_align
            empty_cell.font = Font(name="微软雅黑", size=10, italic=True, color="94a3b8")
            for col_idx in range(1, 10):
                sal_sheet.cell(row=current_row, column=col_idx).border = thin_border
            current_row += 1
            
        # 3. 在右侧展现“当月收支明细流水” (从第 L 列 (第12列) 开始，从第 3 行开始写入以错开标题)
        payment_headers = ["流水工号", "员工姓名", "交易类别", "金额", "交易执行日期", "备注说明"]
        sal_sheet.cell(row=1, column=12, value="月度工资预支与发放流水明细").font = Font(name="微软雅黑", size=12, bold=True, color="0f766e")
        
        for col_idx, h in enumerate(payment_headers, start=12):
            cell = sal_sheet.cell(row=3, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # 获取流水
        payment_data = repo.list_salary_payments(month=month)
        
        pay_row = 4
        for idx, pay in enumerate(payment_data, start=1):
            sal_sheet.cell(row=pay_row, column=12, value=f"PAY-{pay['id']:05d}").alignment = center_align
            sal_sheet.cell(row=pay_row, column=13, value=pay["person_name"]).alignment = center_align
            sal_sheet.cell(row=pay_row, column=14, value=pay["payment_type"]).alignment = center_align
            
            amt_cell = sal_sheet.cell(row=pay_row, column=15, value=pay["amount"])
            amt_cell.alignment = Alignment(horizontal="right", vertical="center")
            amt_cell.number_format = '"¥"#,##0.00'
            if pay["payment_type"] == "预支工资":
                amt_cell.font = Font(name="微软雅黑", size=10, color="d97706")
            else:
                amt_cell.font = Font(name="微软雅黑", size=10, color="10b981")
                
            sal_sheet.cell(row=pay_row, column=16, value=pay["payment_date"]).alignment = center_align
            sal_sheet.cell(row=pay_row, column=17, value=pay["notes"] or "--").alignment = Alignment(horizontal="left", vertical="center")
            
            for col_idx in range(12, 18):
                c = sal_sheet.cell(row=pay_row, column=col_idx)
                c.border = thin_border
                if col_idx != 15:
                    c.font = Font(name="微软雅黑", size=10)
            pay_row += 1
            
        # 5. 自适应列宽
        for col in sal_sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                byte_len = len(val_str.encode("utf-8"))
                char_len = len(val_str)
                approx_len = char_len + (byte_len - char_len) // 2
                if approx_len > max_len:
                    max_len = approx_len
            sal_sheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

    return workbook


def build_contract_workbook(path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "项目合同台账"
    
    # 启用网格线显示
    sheet.views.sheetView[0].showGridLines = True

    # 表头定义
    headers = ["合同ID", "归属项目", "合同名称", "合同分类", "备注", "创建时间"]
    sheet.append(headers)

    # 样式配置
    emerald_color = "0F766E"      # 深邃翠绿表头
    zebra_color = "F0FDF4"        # 浅绿斑马纹
    border_color = "CBD5E1"

    font_header = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="微软雅黑", size=10)
    
    fill_header = PatternFill(start_color=emerald_color, end_color=emerald_color, fill_type="solid")
    fill_zebra = PatternFill(start_color=zebra_color, end_color=zebra_color, fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_side = Side(border_style="thin", color=border_color)
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 设置表头样式
    sheet.row_dimensions[1].height = 28
    for col_idx, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = grid_border

    # 写入数据
    contracts = repo.list_contracts()
    for row_idx, contract in enumerate(contracts, start=2):
        sheet.row_dimensions[row_idx].height = 22
        
        # 数据准备
        row_data = [
            contract["id"],
            contract["project_name"],
            contract["name"],
            contract["contract_type"],
            contract["notes"],
            contract["created_at"][:19] if contract["created_at"] else ""
        ]
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = grid_border
            
            # 斑马纹交替行
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
                
            # 对齐
            if col_idx in (1, 4, 6):
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # 自动调整列宽
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or "")
            byte_len = len(val_str.encode("utf-8"))
            char_len = len(val_str)
            approx_len = char_len + (byte_len - char_len) // 2
            if approx_len > max_len:
                max_len = approx_len
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


