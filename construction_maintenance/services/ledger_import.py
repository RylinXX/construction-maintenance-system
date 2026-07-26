from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
import math
from pathlib import Path, PurePosixPath
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from construction_maintenance.finance import (
    CLASSIFICATION_CONFIDENCES,
    LEDGER_CATEGORY_TREE,
    PAYMENT_STATUSES,
    TRANSACTION_TYPES,
)


DETAIL_HEADERS = (
    "记录编号", "项目名称", "发生日期", "收支类型", "一级分类", "二级分类",
    "事项摘要", "金额（元）", "经办/垫付人", "支付状态", "支付日期",
    "支付/报销说明", "备注（原始用途/购买原因）", "分类置信度",
    "来源文件", "来源工作表", "原始行号",
)
PENDING_HEADERS = (
    "记录类型", "项目名称", "发生日期", "事项/原始用途", "当前一级分类",
    "当前二级分类", "待确认问题", "经办/垫付人", "支付说明", "来源文件",
    "来源工作表", "原始行号",
)


@dataclass(frozen=True)
class LedgerEntry:
    source_record_id: str
    project_name: str
    entry_date: str
    transaction_type: str
    primary_category: str
    secondary_category: str
    summary: str
    amount: float
    handler_name: str
    payment_status: str
    payment_date: str
    payment_notes: str
    original_notes: str
    classification_confidence: str
    source_filename: str
    source_sheet: str
    source_row: int
    review_status: str


@dataclass(frozen=True)
class PendingLedgerItem:
    project_name: str
    item_date: str
    summary: str
    primary_category: str
    secondary_category: str
    handler_name: str
    payment_notes: str
    source_filename: str
    source_sheet: str
    source_row: int
    issue_type: str


@dataclass(frozen=True)
class LedgerImportPreview:
    project_names: tuple[str, ...]
    entries: tuple[LedgerEntry, ...]
    pending_items: tuple[PendingLedgerItem, ...]
    totals: dict[str, float]

    @property
    def project_count(self) -> int:
        return len(self.project_names)


DEMO_PROJECTS = {
    "中央电视总台项目": ("进行中", "中央电视台", "2026-01-01", "2026-12-31", "包含中央电视总台项目资料"),
    "军庄项目": ("进行中", "军庄建设方", "2026-02-01", "2026-12-31", "包含军庄资料"),
    "衙门口项目": ("进行中", "衙门口建设方", "2026-03-01", "2026-12-31", "包含衙门口资料"),
    "老东山项目": ("已完工", "老东山建设方", "2025-05-01", "2026-05-01", "老东山资料-完工"),
    "通州潞城项目": ("进行中", "通州区潞城建设", "2026-04-01", "2026-12-31", "通州潞城项目资料"),
    "内蒙二期项目": ("已完工", "内蒙电力", "2025-06-01", "2026-05-01", "内蒙二期项目-完工"),
    "首师大八里庄项目": ("进行中", "首师大", "2026-04-15", "2026-12-31", "首师大八里庄项目资料"),
    "北理工项目": ("已完工", "北京理工大学", "2025-07-01", "2026-05-01", "北理工项目-完工"),
    "顺义项目": ("进行中", "顺义建设方", "2026-05-01", "2026-12-31", "顺义项目资料"),
    "通州六合工地项目": ("进行中", "通州区六合", "2026-03-10", "2026-12-31", "通州六合工地"),
    "新兴项目": ("进行中", "新兴建设方", "2026-02-15", "2026-12-31", "新兴资料"),
    "梧桐苑项目": ("进行中", "梧桐苑房地产", "2026-01-10", "2026-12-31", "梧桐苑项目资料"),
}
DEMO_VOUCHERS = {
    ("中央电视总台项目", "2026-05-20", "材料费用", 15200.0, "中央电视总台项目采购电缆一批"),
    ("军庄项目", "2026-05-24", "转账凭证", 4800.0, "军庄项目 - 运输运费报销"),
    ("中央电视总台项目", "2026-05-28", "油费", 2400.0, "项目车辆5月油卡充值报销凭证"),
}
DEMO_CONTRACTS = {
    ("中央电视总台项目", "中央电视总台项目劳务分包合同", "劳务合同", "contract_metro_labor.pdf"),
    ("军庄项目", "军庄项目绿化苗木采购合同", "材料商合同", "contract_green_tree.pdf"),
}


def _iso_date(value, *, field: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise ValueError(f"{field}不是有效日期") from exc


def _headers(sheet, expected: tuple[str, ...]) -> dict[str, int]:
    values = [cell.value for cell in sheet[3]]
    positions = {str(value).strip(): index for index, value in enumerate(values) if value}
    missing = [name for name in expected if name not in positions]
    if missing:
        raise ValueError(f"缺少字段: {', '.join(missing)}")
    return positions


def _value(row, positions: dict[str, int], name: str):
    index = positions[name]
    return row[index] if index < len(row) else None


def _category_pairs() -> dict[tuple[str, str], str]:
    return {
        (primary, secondary): scope
        for primary, (scope, secondaries) in LEDGER_CATEGORY_TREE.items()
        for secondary in secondaries
    }


def _parse_workbook(workbook, workbook_name: str):
    if "费用明细" not in workbook.sheetnames or "待确认清单" not in workbook.sheetnames:
        raise ValueError(f"{workbook_name}: 必须包含费用明细和待确认清单")

    pending_sheet = workbook["待确认清单"]
    pending_headers = _headers(pending_sheet, PENDING_HEADERS)
    review_keys: set[tuple[str, str, int]] = set()
    pending_items: list[PendingLedgerItem] = []
    project_names: set[str] = set()
    category_pairs = _category_pairs()
    for excel_row, row in enumerate(
        pending_sheet.iter_rows(min_row=4, values_only=True), start=4
    ):
        record_type = str(_value(row, pending_headers, "记录类型") or "").strip()
        if not record_type:
            continue
        project_name = str(_value(row, pending_headers, "项目名称") or "").strip()
        source_filename = str(_value(row, pending_headers, "来源文件") or "").strip()
        source_sheet = str(_value(row, pending_headers, "来源工作表") or "").strip()
        try:
            source_row = int(_value(row, pending_headers, "原始行号"))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{workbook_name}/待确认清单/{excel_row}: 原始行号无效") from exc
        project_names.add(project_name)
        key = (source_filename, source_sheet, source_row)
        if record_type == "有金额待复核":
            review_keys.add(key)
            continue
        if record_type != "缺少金额":
            raise ValueError(f"{workbook_name}/待确认清单/{excel_row}: 记录类型无效")
        primary = str(_value(row, pending_headers, "当前一级分类") or "").strip()
        secondary = str(_value(row, pending_headers, "当前二级分类") or "").strip()
        if (primary, secondary) not in category_pairs:
            raise ValueError(f"{workbook_name}/待确认清单/{excel_row}: 分类不存在")
        pending_items.append(PendingLedgerItem(
            project_name=project_name,
            item_date=_iso_date(_value(row, pending_headers, "发生日期"), field="发生日期"),
            summary=str(_value(row, pending_headers, "事项/原始用途") or "").strip(),
            primary_category=primary,
            secondary_category=secondary,
            handler_name=str(_value(row, pending_headers, "经办/垫付人") or "").strip(),
            payment_notes=str(_value(row, pending_headers, "支付说明") or "").strip(),
            source_filename=source_filename,
            source_sheet=source_sheet,
            source_row=source_row,
            issue_type=str(_value(row, pending_headers, "待确认问题") or "").strip(),
        ))

    detail_sheet = workbook["费用明细"]
    detail_headers = _headers(detail_sheet, DETAIL_HEADERS)
    entries: list[LedgerEntry] = []
    for excel_row, row in enumerate(
        detail_sheet.iter_rows(min_row=4, values_only=True), start=4
    ):
        record_id = str(_value(row, detail_headers, "记录编号") or "").strip()
        if not record_id:
            continue
        project_name = str(_value(row, detail_headers, "项目名称") or "").strip()
        project_names.add(project_name)
        transaction_type = str(_value(row, detail_headers, "收支类型") or "").strip()
        if transaction_type not in TRANSACTION_TYPES:
            raise ValueError(f"{workbook_name}/费用明细/{excel_row}: 收支类型无效")
        primary = str(_value(row, detail_headers, "一级分类") or "").strip()
        secondary = str(_value(row, detail_headers, "二级分类") or "").strip()
        scope = category_pairs.get((primary, secondary))
        expected_scope = "支出" if transaction_type in {"支出", "冲减支出"} else transaction_type
        if scope != expected_scope:
            raise ValueError(f"{workbook_name}/费用明细/{excel_row}: 分类与收支类型不匹配")
        amount = _value(row, detail_headers, "金额（元）")
        if not isinstance(amount, (int, float)) or not math.isfinite(float(amount)) or amount <= 0:
            raise ValueError(f"{workbook_name}/费用明细/{excel_row}: 金额无效")
        payment_status = str(_value(row, detail_headers, "支付状态") or "").strip()
        if payment_status not in PAYMENT_STATUSES:
            raise ValueError(f"{workbook_name}/费用明细/{excel_row}: 付款状态无效")
        confidence = str(_value(row, detail_headers, "分类置信度") or "").strip()
        if confidence and confidence not in CLASSIFICATION_CONFIDENCES:
            raise ValueError(f"{workbook_name}/费用明细/{excel_row}: 分类置信度无效")
        source_filename = str(_value(row, detail_headers, "来源文件") or "").strip()
        source_sheet = str(_value(row, detail_headers, "来源工作表") or "").strip()
        try:
            source_row = int(_value(row, detail_headers, "原始行号"))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{workbook_name}/费用明细/{excel_row}: 原始行号无效") from exc
        payment_date_value = _value(row, detail_headers, "支付日期")
        entries.append(LedgerEntry(
            source_record_id=record_id,
            project_name=project_name,
            entry_date=_iso_date(_value(row, detail_headers, "发生日期"), field="发生日期"),
            transaction_type=transaction_type,
            primary_category=primary,
            secondary_category=secondary,
            summary=str(_value(row, detail_headers, "事项摘要") or "").strip(),
            amount=float(amount),
            handler_name=str(_value(row, detail_headers, "经办/垫付人") or "").strip(),
            payment_status=payment_status,
            payment_date=_iso_date(payment_date_value, field="支付日期") if payment_date_value else "",
            payment_notes=str(_value(row, detail_headers, "支付/报销说明") or "").strip(),
            original_notes=str(_value(row, detail_headers, "备注（原始用途/购买原因）") or "").strip(),
            classification_confidence=confidence,
            source_filename=source_filename,
            source_sheet=source_sheet,
            source_row=source_row,
            review_status=(
                "待复核" if (source_filename, source_sheet, source_row) in review_keys else "已确认"
            ),
        ))
    if len(project_names) != 1:
        raise ValueError(f"{workbook_name}: 一个账套只能包含一个项目")
    return entries, pending_items, project_names


def parse_ledger_source(path: Path) -> LedgerImportPreview:
    path = Path(path)
    workbook_sources: list[tuple[str, object]] = []
    if path.suffix.lower() == ".xlsx":
        workbook_sources.append((path.name, path))
    elif path.suffix.lower() == ".zip":
        try:
            with ZipFile(path) as archive:
                for member in archive.infolist():
                    member_path = PurePosixPath(member.filename)
                    if (
                        member_path.is_absolute()
                        or ".." in member_path.parts
                        or "\\" in member.filename
                    ):
                        raise ValueError(f"不安全的压缩包路径: {member.filename}")
                    if member.is_dir() or len(member_path.parts) != 1:
                        continue
                    if member_path.suffix.lower() == ".xlsx":
                        workbook_sources.append((member_path.name, BytesIO(archive.read(member))))
        except BadZipFile as exc:
            raise ValueError("压缩包损坏") from exc
    else:
        raise ValueError("仅支持 ZIP 或 XLSX 账套")
    if not workbook_sources:
        raise ValueError("没有找到可导入的 XLSX 账套")

    entries: list[LedgerEntry] = []
    pending_items: list[PendingLedgerItem] = []
    projects: set[str] = set()
    record_ids: set[str] = set()
    for workbook_name, source in workbook_sources:
        workbook = load_workbook(source, read_only=True, data_only=False)
        parsed_entries, parsed_pending, parsed_projects = _parse_workbook(workbook, workbook_name)
        for entry in parsed_entries:
            if entry.source_record_id in record_ids:
                raise ValueError(f"重复记录编号: {entry.source_record_id}")
            record_ids.add(entry.source_record_id)
        entries.extend(parsed_entries)
        pending_items.extend(parsed_pending)
        projects.update(parsed_projects)

    totals = {
        "expense": sum(e.amount for e in entries if e.transaction_type == "支出"),
        "expense_reduction": sum(e.amount for e in entries if e.transaction_type == "冲减支出"),
        "income": sum(e.amount for e in entries if e.transaction_type == "收入"),
        "fund_transfer": sum(e.amount for e in entries if e.transaction_type == "资金往来"),
    }
    totals["net_expense"] = totals["expense"] - totals["expense_reduction"]
    return LedgerImportPreview(
        project_names=tuple(sorted(projects)),
        entries=tuple(entries),
        pending_items=tuple(pending_items),
        totals=totals,
    )


def _delete_verified_demo_projects(db) -> None:
    placeholders = ",".join("?" for _ in DEMO_PROJECTS)
    projects = db.execute(
        f"select * from projects where name in ({placeholders})",
        tuple(DEMO_PROJECTS),
    ).fetchall()
    if not projects:
        return

    matching_rows = []
    for row in projects:
        actual = (
            row["status"], row["owner"], row["start_date"],
            row["end_date"], row["notes"],
        )
        if actual == DEMO_PROJECTS[row["name"]]:
            matching_rows.append(row)
    if not matching_rows:
        return
    if len(projects) != len(DEMO_PROJECTS) or len(matching_rows) != len(DEMO_PROJECTS):
        raise ValueError("演示项目清单与生产库不一致，停止清理")

    project_ids = {int(row["id"]): row["name"] for row in matching_rows}
    id_placeholders = ",".join("?" for _ in project_ids)
    vouchers = db.execute(
        f"""
        select project_id, voucher_date, voucher_type, amount, notes
        from vouchers where project_id in ({id_placeholders})
        """,
        tuple(project_ids),
    ).fetchall()
    voucher_manifest = {
        (
            project_ids[int(row["project_id"])], row["voucher_date"],
            row["voucher_type"], float(row["amount"]), row["notes"],
        )
        for row in vouchers
    }
    contracts = db.execute(
        f"""
        select project_id, name, contract_type, attachment_path
        from contracts where project_id in ({id_placeholders})
        """,
        tuple(project_ids),
    ).fetchall()
    contract_manifest = {
        (
            project_ids[int(row["project_id"])], row["name"],
            row["contract_type"], row["attachment_path"],
        )
        for row in contracts
    }
    if voucher_manifest != DEMO_VOUCHERS or contract_manifest != DEMO_CONTRACTS:
        raise ValueError("演示项目下存在非演示财务数据，停止清理")
    db.execute(
        f"delete from contracts where project_id in ({id_placeholders})",
        tuple(project_ids),
    )
    db.execute(
        f"delete from vouchers where project_id in ({id_placeholders})",
        tuple(project_ids),
    )
    db.execute(
        f"delete from projects where id in ({id_placeholders})",
        tuple(project_ids),
    )


def apply_ledger_import(
    preview: LedgerImportPreview,
    *,
    replace_demo_projects: bool,
    actor_admin_id: int | None,
) -> dict[str, int | float]:
    from construction_maintenance import repositories as repo
    from construction_maintenance.db import get_db

    db = get_db()
    db.execute("begin immediate")
    try:
        if replace_demo_projects:
            _delete_verified_demo_projects(db)
        main_company = db.execute(
            "select id from companies where is_main = 1 order by id limit 1"
        ).fetchone()
        if main_company is None:
            raise ValueError("生产库没有主公司，不能创建项目")
        category_ids = {
            (row["primary_name"], row["secondary_name"]): int(row["id"])
            for row in db.execute(
                """
                select leaf.id, leaf.name as secondary_name,
                       parent.name as primary_name
                from expense_categories leaf
                join expense_categories parent on parent.id = leaf.parent_id
                where leaf.is_active = 1 and parent.is_active = 1
                """
            ).fetchall()
        }
        project_ids: dict[str, int] = {}
        for project_name in preview.project_names:
            existing = db.execute(
                "select id from projects where name = ?", (project_name,)
            ).fetchone()
            if existing:
                project_ids[project_name] = int(existing["id"])
                continue
            project_entries = [
                entry for entry in preview.entries if entry.project_name == project_name
            ]
            project_pending = [
                item for item in preview.pending_items if item.project_name == project_name
            ]
            source_dates = [entry.entry_date for entry in project_entries]
            source_dates.extend(item.item_date for item in project_pending)
            if not source_dates:
                raise ValueError(f"项目 {project_name} 没有可导入的账套记录")
            start_date = min(source_dates)
            end_of_source_period = max(source_dates)
            cursor = db.execute(
                """
                insert into projects (
                  company_id, name, status, owner, start_date, end_date, notes
                ) values (?, ?, '进行中', '', ?, '', ?)
                """,
                (
                    int(main_company["id"]),
                    project_name,
                    start_date,
                    f"标准账套导入；账目期间 {start_date} 至 {end_of_source_period}",
                ),
            )
            project_ids[project_name] = int(cursor.lastrowid)

        inserted_entries = 0
        for entry in preview.entries:
            exists = db.execute(
                "select 1 from vouchers where source_record_id = ?",
                (entry.source_record_id,),
            ).fetchone()
            if exists:
                continue
            category_id = category_ids.get(
                (entry.primary_category, entry.secondary_category)
            )
            if category_id is None:
                raise ValueError(
                    f"分类不存在: {entry.primary_category}/{entry.secondary_category}"
                )
            repo._insert_voucher(db, {
                "project_id": project_ids[entry.project_name],
                "voucher_date": entry.entry_date,
                "transaction_type": entry.transaction_type,
                "category_id": category_id,
                "amount": entry.amount,
                "notes": entry.summary,
                "handler_name": entry.handler_name,
                "payment_status": entry.payment_status,
                "payment_date": entry.payment_date,
                "payment_notes": entry.payment_notes,
                "review_status": entry.review_status,
                "classification_confidence": entry.classification_confidence,
                "source_record_id": entry.source_record_id,
                "source_filename": entry.source_filename,
                "source_sheet": entry.source_sheet,
                "source_row": entry.source_row,
                "original_notes": entry.original_notes,
                "entry_user": "账套导入",
            })
            inserted_entries += 1

        inserted_pending = 0
        for item in preview.pending_items:
            category_id = category_ids.get(
                (item.primary_category, item.secondary_category)
            )
            if category_id is None:
                raise ValueError(
                    f"分类不存在: {item.primary_category}/{item.secondary_category}"
                )
            cursor = db.execute(
                """
                insert or ignore into ledger_pending_items (
                  project_id, item_date, summary, suggested_category_id,
                  handler_name, payment_notes, source_filename, source_sheet,
                  source_row, issue_type
                ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_ids[item.project_name], item.item_date, item.summary,
                    category_id, item.handler_name, item.payment_notes,
                    item.source_filename, item.source_sheet, item.source_row,
                    item.issue_type,
                ),
            )
            inserted_pending += int(cursor.rowcount > 0)
        repo._insert_audit(
            db,
            actor_admin_id=actor_admin_id,
            action="import",
            entity_type="project_ledger",
            entity_id=None,
            details={
                "projects": len(project_ids),
                "entries": inserted_entries,
                "pending_items": inserted_pending,
                "totals": preview.totals,
            },
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {
        "projects": len(project_ids),
        "entries": inserted_entries,
        "pending_items": inserted_pending,
        **preview.totals,
    }
