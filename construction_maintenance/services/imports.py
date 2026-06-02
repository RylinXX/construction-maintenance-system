from __future__ import annotations

import mimetypes
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from construction_maintenance import repositories as repo


def _secure_upload_filename(file: FileStorage) -> str:
    raw_filename = file.filename or "upload"
    filename = secure_filename(raw_filename) or "upload"
    original_suffix = Path(raw_filename).suffix.lower()

    if Path(filename).suffix:
        return filename

    if original_suffix:
        base = filename
        if base.lower() == original_suffix.lstrip("."):
            base = "upload"
        return f"{base}{original_suffix}"

    guessed_suffix = mimetypes.guess_extension(file.mimetype or "")
    if guessed_suffix:
        return f"{filename}{guessed_suffix}"

    return filename


def save_upload(upload_folder: Path, file: FileStorage) -> Path:
    original = _secure_upload_filename(file)
    filename = f"{uuid4().hex}_{original}"
    upload_folder.mkdir(parents=True, exist_ok=True)
    target = upload_folder / filename
    file.save(target)
    return target


def import_attendance_workbook(file_path: Path, month: str) -> dict:
    import calendar

    try:
        workbook = load_workbook(file_path, data_only=True)
    except Exception as exc:
        return {"status": "error", "message": f"无法解析 Excel 文件: {str(exc)}"}

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"status": "error", "message": "Excel 考勤表为空"}

    headers = rows[0]
    if len(headers) < 3 or headers[0] != "姓名" or headers[1] != "身份证号":
        return {"status": "error", "message": "Excel 表头格式非法。必须以【姓名】和【身份证号】开头"}

    try:
        year, month_num = map(int, month.split("-"))
    except (ValueError, TypeError):
        return {"status": "error", "message": "非法的月份格式"}

    _, total_days = calendar.monthrange(year, month_num)

    if len(headers) < 2 + total_days:
        return {
            "status": "error",
            "message": f"Excel 列数不匹配，{month} 月需要至少 {2 + total_days} 列（已检测到 {len(headers)} 列）",
        }

    # 读取当前系统内全体员工用于比对
    people = repo.list_people()
    people_by_id_number = {str(p["id_number"]).strip(): p for p in people}
    people_by_name = {str(p["name"]).strip(): p for p in people}

    import_count = 0
    errors = []

    # 循环读取员工行并进行入库保存
    for idx, row in enumerate(rows[1:], start=2):
        if not row or len(row) < 2:
            continue
        name = str(row[0]).strip() if row[0] is not None else ""
        id_number = str(row[1]).strip() if row[1] is not None else ""

        if not name and not id_number:
            continue

        # 匹配人员
        person = None
        if id_number:
            person = people_by_id_number.get(id_number)
        if not person and name:
            person = people_by_name.get(name)

        if not person:
            errors.append(f"第 {idx} 行：系统内未找到姓名为【{name}】或身份证为【{id_number}】的人员档案")
            continue

        person_id = person["id"]

        # 解析每一天的打卡字符
        for d in range(1, total_days + 1):
            col_idx = 1 + d  # 1日对应下标 2
            val = row[col_idx] if col_idx < len(row) else None
            val_str = str(val).strip() if val is not None else ""

            shift_type = None
            if val_str in ("白", "白班"):
                shift_type = "白班"
            elif val_str in ("夜", "夜班"):
                shift_type = "夜班"
            elif val_str in ("假", "请假"):
                shift_type = "请假"

            date_str = f"{month}-{d:02d}"
            try:
                repo.save_attendance(person_id, date_str, shift_type)
            except Exception as e:
                errors.append(f"第 {idx} 行【{name}】在 {d}日保存考勤失败: {str(e)}")

        import_count += 1

    return {
        "status": "success" if not errors else "partial",
        "import_count": import_count,
        "errors": errors,
    }
