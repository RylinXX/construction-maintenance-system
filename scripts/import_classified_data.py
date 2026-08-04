import os
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# 设置环境变量以便加载正确的配置
os.environ['FLASK_APP'] = 'construction_maintenance.app:create_app()'

from construction_maintenance.app import create_app
from construction_maintenance.db import get_db

app = create_app()

def run_import(files):
    with app.app_context():
        db = get_db()
        cur = db.cursor()

        # 先清理上次导入的这2个项目的数据以重新全量导入
        target_projects = ['军庄项目', '梧桐苑项目']
        for p_name in target_projects:
            cur.execute("SELECT id FROM projects WHERE name = ?", (p_name,))
            res = cur.fetchone()
            if res:
                p_id = res[0]
                cur.execute("DELETE FROM vouchers WHERE project_id = ?", (p_id,))
                cur.execute("DELETE FROM ledger_pending_items WHERE project_id = ?", (p_id,))
                cur.execute("DELETE FROM projects WHERE id = ?", (p_id,))
                print(f"Cleaned up previous import for: {p_name}")

        # 获取默认 company_id
        cur.execute("SELECT id FROM companies LIMIT 1")
        res = cur.fetchone()
        default_company_id = res[0] if res else 1

        # 获取现有的项目映射
        cur.execute("SELECT id, name FROM projects")
        projects = {name: pid for pid, name in cur.fetchall()}
        
        # 获取现有的科目映射
        cur.execute("SELECT id, name, parent_id FROM expense_categories")
        db_cats = cur.fetchall()
        
        pri_id_to_name = {row[0]: row[1] for row in db_cats if row[2] is None}
        categories = {}
        for row in db_cats:
            cat_id, name, parent_id = row
            if parent_id is not None and parent_id in pri_id_to_name:
                pri_name = pri_id_to_name[parent_id]
                categories[(pri_name, name)] = cat_id

        pri_name_to_id = {v: k for k, v in pri_id_to_name.items()}

        total_vouchers = 0
        total_pending = 0

        for file_path in files:
            path = Path(file_path)
            if not path.exists():
                print(f"File not found: {file_path}")
                continue
                
            project_name = path.name.split('-')[0]
            if project_name not in projects:
                cur.execute("INSERT INTO projects (name, company_id) VALUES (?, ?)", (project_name, default_company_id))
                projects[project_name] = cur.lastrowid
                print(f"Created project: {project_name}")
            project_id = projects[project_name]

            print(f"Processing {path.name} (Project: {project_name})")
            wb = load_workbook(path, data_only=True)
            if '分类汇总' not in wb.sheetnames:
                print(f"Sheet '分类汇总' not found in {path.name}")
                continue
                
            ws = wb['分类汇总']
            
            current_date = None

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue # Skip completely empty rows
                
                date_val, summary, pri_cat, sec_cat, amount, payee, pay_notes, cat_status, cat_basis, src_row = row[:10]
                
                if date_val:
                    if isinstance(date_val, datetime):
                        current_date = date_val.strftime('%Y-%m-%d')
                    else:
                        current_date = str(date_val).strip().replace('.', '-')
                        if current_date.startswith('206-'):
                            current_date = '2026-' + current_date[4:]
                
                if not current_date:
                    continue # Skip if no date available at all
                
                summary = str(summary).strip() if summary else ''
                pri_cat = str(pri_cat).strip() if pri_cat else ''
                sec_cat = str(sec_cat).strip() if sec_cat else ''
                payee = str(payee).strip() if payee else ''
                pay_notes = str(pay_notes).strip() if pay_notes else ''
                cat_basis = str(cat_basis).strip() if cat_basis else ''
                
                try:
                    amt = float(amount) if amount is not None else 0.0
                except (ValueError, TypeError):
                    amt = 0.0

                # Ensure category exists
                cat_id = None
                if pri_cat and sec_cat:
                    if (pri_cat, sec_cat) not in categories:
                        if pri_cat not in pri_name_to_id:
                            cur.execute("INSERT INTO expense_categories (name, parent_id) VALUES (?, NULL)", (pri_cat,))
                            pri_name_to_id[pri_cat] = cur.lastrowid
                            print(f"Created new primary category: {pri_cat}")
                        
                        parent_id = pri_name_to_id[pri_cat]
                        cur.execute("SELECT id FROM expense_categories WHERE name=? AND parent_id=?", (sec_cat, parent_id))
                        res = cur.fetchone()
                        if res:
                            categories[(pri_cat, sec_cat)] = res[0]
                        else:
                            cur.execute("INSERT INTO expense_categories (name, parent_id) VALUES (?, ?)", (sec_cat, parent_id))
                            categories[(pri_cat, sec_cat)] = cur.lastrowid
                            print(f"Created new secondary category: {sec_cat} under {pri_cat}")
                    cat_id = categories[(pri_cat, sec_cat)]

                src_row_val = int(src_row) if (src_row and str(src_row).isdigit()) else 0

                if amt > 0:
                    # 有有效金额 -> 录入 vouchers 表
                    review_status = '已确认' if cat_status == '已归类' else '未确认'
                    cur.execute("""
                        INSERT INTO vouchers (
                            project_id, voucher_date, voucher_type, amount, notes, 
                            category_id, handler_name, payment_notes, review_status, classification_confidence,
                            source_filename, source_sheet, source_row, original_notes
                        ) VALUES (?, ?, '支出', ?, ?, ?, ?, ?, ?, ?, ?, '分类汇总', ?, ?)
                    """, (
                        project_id, current_date, amt, summary, cat_id, payee, pay_notes, review_status, cat_basis, path.name, src_row_val, summary
                    ))
                    total_vouchers += 1
                else:
                    # 无金额/0元记录（如台班工时待核算） -> 录入 ledger_pending_items 表
                    issue_type = '缺失金额/待核算金额' if amt == 0 else '未设定分类或金额'
                    cur.execute("""
                        INSERT OR IGNORE INTO ledger_pending_items (
                            project_id, item_date, summary, suggested_category_id,
                            handler_name, payment_notes, source_filename, source_sheet,
                            source_row, issue_type
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, '分类汇总', ?, ?)
                    """, (
                        project_id, current_date, summary, cat_id, payee, pay_notes, path.name, src_row_val, issue_type
                    ))
                    total_pending += 1

        db.commit()
        print(f"Full import completed! Successfully inserted {total_vouchers} vouchers and {total_pending} pending ledger items.")

if __name__ == '__main__':
    files = sys.argv[1:]
    run_import(files)
